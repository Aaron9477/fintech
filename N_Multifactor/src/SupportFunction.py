# -*- coding: utf-8 -*-
"""
Created on Mon Mar  7 21:41:51 2022
#各种辅助函数的集合
@author: Hao Li
"""

# 运行系统命令，安装WindPy包，若已安装，可不运行
# !D:\Anaconda3\python.exe C:\Wind\Wind.NET.Client\WindNET\bin\installWindPy.py C:\wind\wind.net.client\windnet
import matplotlib.pyplot as plt
import quantstats as qt
import sqlite3
import pandas as pd
import numpy as np
from WindPy import w
import empyrical as em
import time
import os
from functools import reduce
from itertools import product
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# %% 将数据连接到 DB Browser, 输入为下载所需要的sql语句
def Obtain_DBData(sql_str, path):
    conn = sqlite3.connect(path)
    data = pd.read_sql(sql_str, conn)
    return data


# %% 计时装饰器
def timeit_test(func):
    def wrapper(*args, **kwargs):
        start = time.perf_counter()
        func(*args, **kwargs)
        elapsed = time.perf_counter() - start
        print('Time used:{:.2f}s'.format(elapsed))

    return wrapper


# %% 两组序列的排列组合
def ExpandFundList(fund_list, date_list, colnames):
    aa = pd.DataFrame()
    for fund in fund_list:
        for date in date_list:
            bb = pd.Series([fund, date])
            aa = aa.append(bb, ignore_index=True)
    aa.columns = colnames
    return aa


# ExpandFundList(['a','b','c'],[1,2,3],colnames=['col1','col2'])

# %% 权重组合的排列组合
def ExpandWeight(weight_source, colnames):
    fn = lambda x, code='|': reduce(lambda x, y: [str(i) + code + str(j) for i in x for j in y], x)
    temp0 = list(map(lambda x: list(map(float, str.split(x, "+"))), fn(weight_source, '+')))
    temp = pd.DataFrame(temp0)[pd.DataFrame(temp0).sum(axis=1) == 1].reset_index(drop=True)
    temp.columns = colnames
    # 方法二
    # w = np.linspace(0, 1, 11)
    # perms = pd.DataFrame(product(w, repeat=8))
    # perms.loc[perms.sum(axis=1) == 1].reset_index(drop=True)
    return temp


# 最多支持7组权重
# ExpandWeight(weight_source=[np.arange(0, 1.1, 0.1)] * 3, colnames=["w1", "w2", "w3", "w4", 'w5', 'w6'])
# ExpandWeight(weight_source=[np.arange(0, 1.1, 0.1),np.arange(0, 1.1, 0.2), np.arange(0, 1.1, 0.4)], colnames=["w1", "w2", "w3"])

# %% 字典输出结果导出到Excel文件（目前只支持两层字典，第二层可以是列表或者Dataframe
# pic_include列示表格sheet名称和对应的存放于工作目录的图片文件名
def output_to_excel(output_dict="", filename="", resetcol=True, pic_include=dict()):
    writer = pd.ExcelWriter(filename)
    for key in output_dict:
        if type(output_dict[key]) == list:
            output_dict[key] = pd.DataFrame(output_dict[key])
        output_dict[key].to_excel(writer, sheet_name=key, float_format='%.6f')

    if pic_include:  # 如果字典不为空
        for key,value in pic_include.items():
            worksheet = writer.sheets[key]
            worksheet.insert_image('H4', value)
    writer.save()
    writer.close()

    if resetcol == True:
        reset_col(filename)  # 重设列宽
    print('结果输出完毕！')


# %% 导出的Excel重设列宽
def reset_col(filename):
    wb = load_workbook(filename)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        df = pd.read_excel(filename, sheet).fillna('-')
        df.loc[len(df)] = list(df.columns)  # 把标题行附加到最后一行
        for col in df.columns:
            index = list(df.columns).index(col)  # 列序号
            letter = get_column_letter(index + 1)  # 列字母
            collen = df[col].apply(lambda x: len(str(x).encode())).max()  # 获取这一列长度的最大值 当然也可以用min获取最小值 mean获取平均值
            ws.column_dimensions[letter].width = collen * 1.2 + 4  # 也就是列宽为最大长度*1.2 可以自己调整
    wb.save(filename)


# %% 降低物理内存占用
def reduce_mem_usage(df, verbose=True):
    numerics = ['int16', 'int32', 'int64', 'float16', 'float32', 'float64']
    start_mem = df.memory_usage().sum() / 1024**2
    for col in df.columns:
        col_type = df[col].dtypes
        if col_type in numerics:
            c_min = df[col].min()
            c_max = df[col].max()
            if str(col_type)[:3] == 'int':
                if c_min > np.iinfo(np.int8).min and c_max < np.iinfo(np.int8).max:
                    df[col] = df[col].astype(np.int8)
                elif c_min > np.iinfo(np.int16).min and c_max < np.iinfo(np.int16).max:
                    df[col] = df[col].astype(np.int16)
                elif c_min > np.iinfo(np.int32).min and c_max < np.iinfo(np.int32).max:
                    df[col] = df[col].astype(np.int32)
                elif c_min > np.iinfo(np.int64).min and c_max < np.iinfo(np.int64).max:
                    df[col] = df[col].astype(np.int64)
            else:
                if c_min > np.finfo(np.float16).min and c_max < np.finfo(np.float16).max:
                    df[col] = df[col].astype(np.float16)
                elif c_min > np.finfo(np.float32).min and c_max < np.finfo(np.float32).max:
                    df[col] = df[col].astype(np.float32)
                else:
                    df[col] = df[col].astype(np.float64)
    end_mem = df.memory_usage().sum() / 1024**2
    if verbose: print('Mem. usage decreased to {:5.2f} Mb ({:.1f}% reduction)'.format(end_mem, 100 * (start_mem - end_mem) / start_mem))
    return df

# %% 长数据有缺失值时进行填充（沿用最近交易日的非空价格）
def long_asset_df_fillna(asset_df):
    assetdf_col = asset_df.columns
    asset_df = asset_df.pivot(index=assetdf_col[0], columns=assetdf_col[1], values=assetdf_col[2]).fillna(
        method='ffill')
    asset_df = asset_df.reset_index().melt(id_vars=assetdf_col[0], var_name=assetdf_col[1], value_name=assetdf_col[2])
    asset_df.dropna(inplace=True)
    return asset_df


# asset_df = fetch_asset_series(stock_code="", index_code=index_list, fund_code=fund_list,
#                               start_date=db_start_date, end_date=db_end_date, data_mode='DB',
#                               censor_rebase=False, to_wide=False)
# asset_df = asset_df[['TRADE_DT', 'S_INFO_WINDCODE', 'S_DQ_ADJCLOSE']] # [329961 rows x 3 columns]
# asset_df1 = long_asset_df_fillna(asset_df) # [334016 rows x 3 columns]
#        TRADE_DT F_INFO_WINDCODE  F_NAV_ADJUSTED
# 0    2009-01-05       270009.OF        1.118000
# 1    2009-01-06       270009.OF        1.120000
# 2    2009-01-07       270009.OF        1.121000
